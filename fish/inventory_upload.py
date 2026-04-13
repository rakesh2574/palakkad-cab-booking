"""
Parse a daily inventory .xlsx upload and insert rows into fish_inventory.

Expected columns (header row): Fish Name | Available Kg | Price Per Kg | Notes
"""

from datetime import datetime, timezone, timedelta
from openpyxl import load_workbook
from . import database as fdb


# Kerala is IST (UTC+5:30). Railway runs in UTC, so we fix the timezone here.
IST = timezone(timedelta(hours=5, minutes=30))


def today_ist_iso():
    return datetime.now(IST).date().isoformat()


REQUIRED_COLS = ["fish name", "available kg", "price per kg"]
OPTIONAL_COLS = ["notes"]


def _norm(s):
    return str(s or "").strip().lower()


def parse_and_load(xlsx_path, inventory_date=None):
    if inventory_date is None:
        inventory_date = today_ist_iso()

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"status": "error", "message": "Empty spreadsheet"}

    header = [_norm(c) for c in rows[0]]
    col_idx = {}
    for col in REQUIRED_COLS + OPTIONAL_COLS:
        if col in header:
            col_idx[col] = header.index(col)

    missing = [c for c in REQUIRED_COLS if c not in col_idx]
    if missing:
        return {
            "status": "error",
            "message": f"Missing required columns: {', '.join(missing)}. "
                       f"Expected: {', '.join(REQUIRED_COLS + OPTIONAL_COLS)}",
        }

    inserted, errors = 0, []
    for i, row in enumerate(rows[1:], start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        try:
            fish_name = str(row[col_idx["fish name"]] or "").strip()
            if not fish_name:
                continue
            available_kg = float(row[col_idx["available kg"]] or 0)
            price = float(row[col_idx["price per kg"]] or 0)
            notes = str(row[col_idx["notes"]] or "").strip() if "notes" in col_idx else ""
            if available_kg < 0 or price < 0:
                raise ValueError("negative value")
            fdb.upsert_inventory(inventory_date, fish_name, available_kg, price, notes)
            inserted += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")

    return {"status": "ok", "inventory_date": inventory_date,
            "inserted": inserted, "errors": errors}
